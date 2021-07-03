import requests
from bs4 import BeautifulSoup
from urllib.request import urlopen

import pytesseract as pyt
import pyautogui as pg

import tkinter as tk
import tkinter.filedialog as fd
from tkinter import messagebox
from tkinter import *

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

import os
import cv2
import pandas as pd
import numpy as np
import datetime

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

'''
파일 저장 함수
url 읽어오기
'''
def get_html(url):
    _html = ""
    resp = requests.get(url)
    if resp.status_code == 200:
        _html = resp.text

    return _html


class UI:
    def __init__(self):        
        self.name = []   # 길드원 이름 저장
        self.save_m = [] # 주간미션 저장
        self.save_u = [] # 지하수로 저장
        self.save_f = [] # 플래그 저장
                
        self.master_name = [] # 길드 마스터 저장
        self.guild_url = []   # 길드 url 저장
        self.guild_list = []  # 길드 이름 저장
                
        self.s_count = 11 # 스크린샷 카운트 버튼
        self.date = datetime.datetime.today().strftime("%Y%m%d") # 프로그램 실행 날짜
                        
        '''
        폴더 생성
        '''
        if os.path.isdir('./'+self.date) != 1:
            os.mkdir('./'+self.date)
            os.mkdir('./'+self.date +'/save')
            os.mkdir('./'+self.date +'/create_image')
            os.mkdir('./'+self.date +'/create_image/flag')
            os.mkdir('./'+self.date +'/create_image/mission')
            os.mkdir('./'+self.date +'/create_image/under')

        '''
        제목 / 크기 / 배경 / 폰트설정
        '''
        self.main_window = tk.Tk()
        self.main_window.title("원기... api 공개해줘")
        self.main_window.geometry("801x420")
        self.main_window.resizable(False, False)        
        label_image = tk.PhotoImage(file = './file/bg.png')
        bg_label = tk.Label(self.main_window, image=label_image)
        bg_label.place(x=0, y=0)
                            
        font = "메이플스토리", 12, "bold"
        font2 = "메이플스토리", 9, "bold"
        
        '''
        진행상황 텍스트
        '''
        self.label = tk.Label(text="Progress ...", font=font)
        self.label.place(x=30, y=20)
        self.progress = tk.Text(self.main_window, width=24, height=23)
        self.progress.configure(font=("Arial", 9, "bold"))
        self.progress.place(x=30, y=50)
        self.text_index = 1.0      
        
        '''
        스크린샷 버튼
        '''
        self.label_stat = tk.Label(self.main_window, text = "[1] 스크린샷 저장", font=font, bg="pink")
        self.label_stat.place(x=250, y=50)        
        btn_stat = tk.Button(self.main_window, text="Screenshot", font=font, bg="pink", overrelief="solid", command=self.call_Screenshot, repeatdelay=1000)
        btn_stat.place(x=250, y=90)
        
        '''
        주간미션, 수로, 플래그 점수 읽기 버튼
        '''
        self.label_create = tk.Label(self.main_window, text = "[2] 사진 분석", font=font, bg="pink")
        self.label_create.place(x=410, y=50)
        btn_create = tk.Button(self.main_window, text="Check", font=font, bg="pink", overrelief="solid", command=self.call_Find_location, repeatdelay=1000)
        btn_create.place(x=410, y=90)
                
        '''
        파일 저장 버튼
        '''
        self.label_save = tk.Label(self.main_window, text = "[3] 변환결과저장", font=font, bg="pink")
        self.label_save.place(x=550, y=50)
        
        self.label_guild = tk.Label(self.main_window, text = "Guild", font=font, bg="pink")
        self.label_guild.place(x=560, y=90)
        self.text_guild = tk.Entry(self.main_window, width=15, textvariable=str)
        self.text_guild.insert(0, '')
        self.text_guild.place(x=630, y=90)
        
        self.label_master = tk.Label(self.main_window, text = "Master", font=font, bg="pink")
        self.label_master.place(x=550, y=120)
        self.text_master = tk.Entry(self.main_window, width=15, textvariable=str)
        self.text_master.insert(0, '')
        self.text_master.place(x=630, y=120)
        
        btn_save = tk.Button(self.main_window, text="Save excel", font=font, bg="pink", overrelief="solid", command=self.call_save_excel, repeatdelay=1000)
        btn_save.place(x=550, y=160)                          

        '''
        파일 분석 버튼        
        '''
        self.label_analysis = tk.Label(self.main_window, text = "[4] 파일분석", font=font, bg="pink")
        self.label_analysis.place(x=250, y=210)        
        btn_analysis = tk.Button(self.main_window, text="Analysis", font=font, bg="pink", overrelief="solid", command=self.call_file_analysis, repeatdelay=1000)
        btn_analysis.place(x=250, y=250)
        
        '''
        주간미션 체크박스
        '''
        self.missionCheckVal=IntVar()
        self.missionCheckbox=Checkbutton(self.main_window,text="주간미션",font=font2, variable=self.missionCheckVal, command = self.call_check_mission)
        self.missionCheckbox.place(x=250, y=300)
        self.missiontext = tk.Entry(self.main_window, width=5, textvariable=str)
        self.missiontext.configure(state='disabled')
        self.missiontext.place(x=330, y=300)

        '''
        지하수로 체크박스
        '''
        self.underCheckVar=IntVar()
        self.underCheckbox=Checkbutton(self.main_window,text="지하수로", font=font2, variable=self.underCheckVar, command = self.call_check_under)
        self.underCheckbox.place(x=250, y=330)        
        self.undertext = tk.Entry(self.main_window, width=5, textvariable=str)
        self.undertext.configure(state='disabled')
        self.undertext.place(x=330, y=330)

        '''
        플래그 체크박스
        '''
        self.flagCheckVar=IntVar()
        self.flagCheckbox=Checkbutton(self.main_window, text="플 래 그", font=font2, variable=self.flagCheckVar, command = self.call_check_flag)
        self.flagCheckbox.place(x=250, y=360)
        self.flagtext = tk.Entry(self.main_window, width=5, textvariable=str)
        self.flagtext.configure(state='disabled')
        self.flagtext.place(x=330, y=360)

        '''
        종료 버튼
        '''
        self.label_exit = tk.Label(self.main_window, text = "[5] 종료", font=font, bg="pink")
        self.label_exit.place(x=410, y=210)        
        close_window = tk.Button(self.main_window, text="Exit", font=font, bg="pink", overrelief="solid", command=self.exit, repeatdelay=1000)        
        close_window.place(x=410, y=250)            
        self.main_window.mainloop()  
    
    def call_check_mission(self):
        if self.missionCheckVal.get() == 1:
            self.missiontext.configure(state='normal')
        if self.missionCheckVal.get() == 0:            
            self.missiontext.delete(0,END)
            self.missiontext.configure(state='disabled')
        
    def call_check_under(self):
        if self.underCheckVar.get() == 1:            
            self.undertext.configure(state='normal')
        if self.underCheckVar.get() == 0:            
            self.undertext.delete(0,END)
            self.undertext.configure(state='disabled') 

    def call_check_flag(self):
        if self.flagCheckVar.get() == 1:            
            self.flagtext.configure(state='normal')
        if self.flagCheckVar.get() == 0:            
            self.flagtext.delete(0,END)
            self.flagtext.configure(state='disabled') 
                               
    '''
    스크린샷
    '''
    def call_Screenshot(self):        
        mission = pg.locateOnScreen('./file/g_mission.png')
        mission = pg.screenshot('./'+self.date+'/save/mission_im'+str(self.s_count)+'.png', region=(mission.left, mission.top+30, mission.width, mission.height+395))
        
        under = pg.locateOnScreen('./file/g_under.png') 
        under = pg.screenshot('./'+self.date+'/save/under_im'+str(self.s_count)+'.png', region=(under.left, under.top+30, under.width, under.height+395))
        
        flag  = pg.locateOnScreen('./file/g_flag.png')                        
        flag  = pg.screenshot('./'+self.date+'/save/flag_im'+str(self.s_count)+'.png', region=(flag.left, flag.top+30, flag.width, flag.height+395))
        
        self.progress.insert(self.text_index, 'save my_screesnshot'+str(self.s_count)+'.png\n')
        self.text_index = self.text_index + 1.0    
        self.s_count+=1

    '''
    길드 창에서 주간미션 / 지하수로 / 플래그레이스 점수를 읽어옴
    '''
    def call_Find_location(self):
        self.progress.delete(1.0, END)
        
        for currentdir, dirs, files in os.walk('./'+self.date+'/save/'):
            for file in files :
                self.name.append(file)
                
        for i in range(0, int(len(self.name)/3)):
            self.progress.insert(self.text_index, str(i+1) + 'th image Conversion...\n')
            self.text_index = self.text_index + 1.0
            self.main_window.update()
            
            
            '''            
            스크린샷에서 점수 부분을 17부분으로 잘라서 저장
            '''
            mission_img  = cv2.imread('./'+self.date+'/save/mission_im'+str(i+11)+'.png')
            under_img = cv2.imread('./'+self.date+'/save/under_im'+str(i+11)+'.png')
            flag_img  = cv2.imread('./'+self.date+'/save/flag_im'+str(i+11)+'.png')
            
            _x, _y, _w, _h = 14, 9, 46, 24
            
            for j in range(0, 17):
                cv2.imwrite('./'+self.date+'/create_image/mission/mission_'+ (str(10+j+(17*i))) + '.png', mission_img[_y:_h, _x:_w])
                cv2.imwrite('./'+self.date+'/create_image/under/under_'+ (str(10+j+(17*i))) + '.png', under_img[_y:_h, _x:_w])
                cv2.imwrite('./'+self.date+'/create_image/flag/flag_'+ (str(10+j+(17*i))) + '.png', flag_img[_y:_h, _x:_w])
                _y = _y+24
                _h = _h+24
                self.main_window.update()
                
                
            '''
            17개로 저장했던 사진 검정색 배경으로 마스킹
            '''            
            for j in range(0, 17): 
                m_img  = cv2.imread('./'+self.date+'/create_image/mission/mission_'+str(10+j+(17*i))+'.png', cv2.IMREAD_GRAYSCALE)
                m_lower=np.percentile(m_img, 5)
                m_upper=np.percentile(m_img,50)
                m_black = cv2.normalize(m_img, m_img, -m_lower, 255+m_upper, cv2.NORM_MINMAX)# 저장
                cv2.imwrite('./'+self.date+'/create_image/mission/mission_'+(str(10+j+(17*i)))+'.png', m_black)
                self.main_window.update()

                u_img = cv2.imread('./'+self.date+'/create_image/under/under_'+str(10+j+(17*i))+'.png', cv2.IMREAD_GRAYSCALE)
                u_lower=np.percentile(u_img, 5)
                u_upper=np.percentile(u_img,50)
                u_black = cv2.normalize(u_img, u_img, -u_lower, 255+u_upper, cv2.NORM_MINMAX)# 저장
                cv2.imwrite('./create_image/under/under_'+(str(10+j+(17*i)))+'.png', u_black)                  
                self.main_window.update()

                f_img  = cv2.imread('./'+self.date+'/create_image/flag/flag_'+str(10+j+(17*i))+'.png', cv2.IMREAD_GRAYSCALE)
                f_lower=np.percentile(f_img, 5)
                f_upper=np.percentile(f_img,50)
                f_black = cv2.normalize(f_img, f_img, -f_lower, 255+f_upper, cv2.NORM_MINMAX)# 저장
                cv2.imwrite('./'+self.date+'/create_image/flag/flag_'+(str(10+j+(17*i)))+'.png', f_black)
                self.main_window.update()
            
            '''
            마스킹한 사진을 숫자 리스트로 저장함    
            '''
            for k in range(0, 17):
                self.save_m.append(str(pyt.image_to_string('./'+self.date+'/create_image/mission/mission_'+str(10+k+(17*i))+'.png', 
                                              lang = None, config='--psm 10 -c tessedit_char_whitelist=0123456789')))
                self.save_u.append(str(pyt.image_to_string('./'+self.date+'/create_image/under/under_'+str(10+k+(17*i))+'.png', 
                                              lang = None, config='--psm 10 -c tessedit_char_whitelist=0123456789')))
                self.save_f.append(str(pyt.image_to_string('./'+self.date+'/create_image/flag/flag_'+str(10+k+(17*i))+'.png', 
                                              lang = None, config='--psm 10 -c tessedit_char_whitelist=0123456789')))
                self.main_window.update()

            cv2.waitKey(0)
            cv2.destroyAllWindows()  
            
        self.progress.insert(self.text_index, 'image Conversion fin...\n')
        self.text_index = self.text_index + 1.0
        self.main_window.update()
        

    '''
    파일저장 함수
    길드 검색
    '''
    def parse_maple_guild(self, guild, master):
        url = "https://maplestory.nexon.com/Ranking/World/Guild?t=1&n=" + guild
        rank_html = get_html(url)
        soup = BeautifulSoup(rank_html, 'html.parser')
        maple_rank = soup.find("table", {"class": "rank_table2"}).find_all("td", {"class": "left"})

        for index in maple_rank:
            info_soup = index.find("a")
            _url = info_soup["href"]
            _text = info_soup.text.split(".")
            _num = _text[0]

            self.master_name.append(_num)
            self.guild_url.append(_url)

        for i in range(0, int(len(self.master_name)/2)):
            if self.master_name[1+i*2] == master:
                self.guild_url.append(self.guild_url[i*2])

    '''
    파일저장 함수
    길드원 닉네임 리스트 오름차순으로 가져옴
    '''    
    def parse_guild_list(self):        
        url = "https://maplestory.nexon.com/" + self.guild_url[len(self.guild_url)-1]
        member = len(self.save_f)
        
        if member % 20 == 0:
            for i in range(0, int(member/20)):            
                rank_html = get_html(url + "&orderby=0&page=" + str(i+1))
                soup = BeautifulSoup(rank_html, 'html.parser')        
                g_index = soup.find("table", {"class":"rank_table"}).find_all("td", {"class":"left"})

                for index in g_index:
                    info_soup = index.find("a")
                    _text = info_soup.text.split(".")
                    _name = _text[0]
                    self.guild_list.append(_name)
        else:
            for j in range(0, int(member/20)+1):                        
                rank_html = get_html(url + "&orderby=0&page=" + str(j+1))
                soup = BeautifulSoup(rank_html, 'html.parser')        
                g_index = soup.find("table", {"class":"rank_table"}).find_all("td", {"class":"left"})

                for index in g_index:
                    info_soup = index.find("a")
                    _text = info_soup.text.split(".")
                    _name = _text[0]
                    self.guild_list.append(_name)

        self.guild_list.sort()
        
    '''
    파일저장 함수
    '''
    def call_save_list(self):        
        self.progress.insert(self.text_index, 'save list ...\n')
        self.text_index = self.text_index + 1.0
        self.main_window.update()
        
        for z in range(0, len(self.save_m)):
            try:
                if self.save_m[z][0] == '\x0c':
                    self.save_m[z] = 0                
                elif len(self.save_m[z]) == 3:
                    self.save_m[z] = int(self.save_m[z][0])
            except TypeError:
                pass

        for z in range(0, len(self.save_u)):
            try:
                if self.save_u[z][0] == '\x0c':
                    self.save_u[z] = 0
                elif len(self.save_u[z]) == 6:
                    self.save_u[z] = int(self.save_u[z][0])*1000 + int(self.save_u[z][1])*100 + int(self.save_u[z][2])*10 + int(self.save_u[z][3])
                elif len(self.save_u[z]) == 5:
                    self.save_u[z] = int(self.save_u[z][0])*100 + int(self.save_u[z][1])*10 + int(self.save_u[z][2])
                elif len(self.save_u[z]) == 4:
                    self.save_u[z] = int(self.save_u[z][0])*10 + int(self.save_u[z][1])
                elif len(self.save_u[z]) == 3:
                    self.save_u[z] = int(self.save_u[z][0])
            except TypeError:
                    pass

        for z in range(0, len(self.save_f)):
            try:
                if self.save_f[z][0] == '\x0c':
                    self.save_f[z] = 0
                elif len(self.save_f[z]) == 6:
                    self.save_f[z] = int(self.save_f[z][0])*1000 + int(self.save_f[z][1])*100 + int(self.save_f[z][2])*10 + int(self.save_f[z][3])
                elif len(self.save_f[z]) == 5:
                    self.save_f[z] = int(self.save_f[z][0])*100 + int(self.save_f[z][1])*10 + int(self.save_f[z][2])
                elif len(self.save_f[z]) == 4:
                    self.save_f[z] = int(self.save_f[z][0])*10 + int(self.save_f[z][1])
                elif len(self.save_f[z]) == 3:
                    self.save_f[z] = int(self.save_f[z][0])
            except TypeError:
                pass
            
        self.progress.insert(self.text_index, 'save list fin...\n')
        self.text_index = self.text_index + 1.0
        self.main_window.update()
        
    '''
    파일 저장
    '''
    def call_save_excel(self):
        self.progress.delete(1.0, END)
        self.progress.insert(self.text_index, 'save excel ...\n')
        self.text_index = self.text_index + 1.0
        self.main_window.update()
        
        guild = self.text_guild.get()
        master = self.text_master.get()    
        
        self.call_save_list()        
        self.parse_maple_guild(guild, master)        
        self.parse_guild_list()
                
        make_list = pd.DataFrame(list(zip(self.guild_list, self.save_m, self.save_u, self.save_f)), columns=['이름', '주간미션', '지하수로', '플래그'])        
        new_base_dir = './'+self.date+'/'
        new_file_name = '길드원 리스트.xlsx'
        new_file_dir = os.path.join(new_base_dir, new_file_name)
        
        make_list.to_excel(new_file_dir,
                           na_rep='NaN',
                           header=True,
                           index=False,
                           startrow=0,
                           startcol=0)

        blank = pd.DataFrame(index=range(0,0), columns=['이름', ' '])

        for naming in make_list['이름']:
            blank = blank.append(pd.DataFrame([[naming,' ']], columns=['이름', '']), ignore_index=True)
            blank = blank.append(pd.DataFrame([[naming + '점수',' ']], columns=['이름', '']), ignore_index=True)

        fill = blank.T.fillna(' ')

        for count in range(0, int(len(blank)/2)):
            fill[count*2][1] = fill[count*2][0]

        new_base_dir2 = './'+self.date+'/'
        new_file_name2 = '여기에 부캐들 정렬할것.xlsx' # 여기를 파일.. 불러오는걸로 아 화장실
        new_file_dir2 = os.path.join(new_base_dir2, new_file_name2)

        fill.to_excel(new_file_dir2,
                      na_rep='NaN',
                      header=False,
                      index=False,
                      startrow=0,
                      startcol=0)  
        
        self.progress.insert(self.text_index, 'save excel fin...\n')
        self.text_index = self.text_index + 1.0
        self.main_window.update()
    
    '''
    파일 분석
    '''
    def call_file_analysis(self):                
        self.progress.delete(1.0, END)
        self.progress.insert(self.text_index, 'analysis ...\n')
        self.text_index = self.text_index + 1.0
        self.main_window.update()        
        
        '''
        엑셀 읽어오기
        '''        
        date = self.date        
        msg1 = messagebox.askokcancel("선택해", "분석할 파일이 있는 폴더를 선택하세요")
        if msg1:
            path = fd.askdirectory(initialdir = "./", title = "수정한 파일이 있는 폴더 선택")
            date = path[path.find("2021"):]
            read_guild_list1 = pd.read_excel(date+'/길드원 리스트.xlsx').fillna(' ')
            read_guild_list2 = pd.read_excel(date+'/여기에 부캐들 정렬할것.xlsx').fillna(' ')

            '''
            unnamed 제거
            '''
            Unnamed = []
            for word in read_guild_list2:
                if word.find("Unnamed") == 0:
                    Unnamed.append(word)
            read_guild_list2.drop(Unnamed, axis=1, inplace=True)

            '''
            각 index 읽어오기
            '''
            naming = []
            for name in range(0, int(len(read_guild_list2.columns)/2)):
                naming.append(read_guild_list2.columns[name*2])

            '''
            점수입력
            '''        
            for data1 in range(0, len(read_guild_list1)):
                for data2 in range(0, len(read_guild_list2)):
                    for data3 in range(0, len(naming)):
                        if read_guild_list1['이름'][data1] == read_guild_list2[naming[data3]][data2]:
                            read_guild_list1['이름']=read_guild_list1['이름'].drop(data1).fillna(' ')
                            read_guild_list2[naming[data3]+'점수'][data2] = str(read_guild_list1['주간미션'][data1]), str(read_guild_list1['지하수로'][data1]), str(read_guild_list1['플래그'][data1])
                            break     

            loc = len(read_guild_list2)

            '''
            주간미션, 지하수로, 플래그 점수 저장
            '''
            score = []
            if self.missionCheckVal.get() == 1:
                score.append(int(self.missiontext.get()))
                score.append(int(loc))
            if self.underCheckVar.get() == 1:        
                score.append(int(self.undertext.get()))
                score.append(int(loc+1))
            if self.flagCheckVar.get() == 1:         
                score.append(int(self.flagtext.get()))
                score.append(int(loc+2))
            
            data4 = {' ' : ' '}
            read_guild_list2 = read_guild_list2.append(data4, ignore_index=True).fillna(' ') # 주간미션용
            read_guild_list2 = read_guild_list2.append(data4, ignore_index=True).fillna(' ') # 지하수로용
            read_guild_list2 = read_guild_list2.append(data4, ignore_index=True).fillna(' ') # 플래그용
            read_guild_list2 = read_guild_list2.append(data4, ignore_index=True).fillna(' ') # 노블여부용
            read_guild_list2 = read_guild_list2.append(data4, ignore_index=True).fillna(' ') # 노블여부용
            read_guild_list2 = read_guild_list2.append(data4, ignore_index=True).fillna(' ') # 빈칸용  

            fm_sum = 0
            fu_sum = 0
            ff_sum = 0

            for data1 in range(0, len(naming)):
                m_sum = 0
                u_sum = 0
                f_sum = 0

                for data2 in range(0, len(read_guild_list2)):
                    try:
                        m_sum = m_sum + int(read_guild_list2[naming[data1]+'점수'][data2][0])
                        u_sum = u_sum + int(read_guild_list2[naming[data1]+'점수'][data2][1])
                        f_sum = f_sum + int(read_guild_list2[naming[data1]+'점수'][data2][2])

                        fm_sum = fm_sum + int(read_guild_list2[naming[data1]+'점수'][data2][0])
                        fu_sum = fu_sum + int(read_guild_list2[naming[data1]+'점수'][data2][1])
                        ff_sum = ff_sum + int(read_guild_list2[naming[data1]+'점수'][data2][2])

                    except ValueError:
                        pass

                read_guild_list2[naming[data1]].loc[loc]=naming[data1]+'주간미션'
                read_guild_list2[naming[data1]+'점수'].loc[loc]=int(m_sum) 

                read_guild_list2[naming[data1]].loc[loc+1]=naming[data1]+'지하수로'
                read_guild_list2[naming[data1]+'점수'].loc[loc+1]=int(u_sum)

                read_guild_list2[naming[data1]].loc[loc+2]=naming[data1]+'플래그'
                read_guild_list2[naming[data1]+'점수'].loc[loc+2]=int(f_sum)                     
                read_guild_list2[naming[data1]].loc[loc+3]=naming[data1]+'노블'                        

                '''
                주간미션, 지하수로, 플래그 점수 분석
                '''
                if len(score) == 6:
                    if (((read_guild_list2[naming[data1]+'점수'][score[1]] >= score[0]) and
                        (read_guild_list2[naming[data1]+'점수'][score[3]] >= score[2]) and 
                        (read_guild_list2[naming[data1]+'점수'][score[5]]>= score[4])) != 1) :
                        read_guild_list2[naming[data1]].loc[loc+4]='너 못써~'

                if len(score) == 4:
                    if (((read_guild_list2[naming[data1]+'점수'][score[1]] >= score[0]) and
                        (read_guild_list2[naming[data1]+'점수'][score[3]] >= score[2])) != 1):
                        read_guild_list2[naming[data1]].loc[loc+4]='너 못써~'

                if len(score) == 2:
                    if (((read_guild_list2[naming[data1]+'점수'][score[1]] >= score[0])) != 1):
                        read_guild_list2[naming[data1]].loc[loc+4]='너 못써~'            

            for data1 in range(0, len(naming)):    
                m = (str("%.f%%" % (int(read_guild_list2[naming[data1]+'점수'][loc]) / int(fm_sum) * 100.0)))
                u = (str("%.f%%" % (int(read_guild_list2[naming[data1]+'점수'][loc+1]) / int(fu_sum) * 100.0)))
                f = (str("%.f%%" % (int(read_guild_list2[naming[data1]+'점수'][loc+2]) / int(ff_sum) * 100.0)))

                read_guild_list2[naming[data1]+'점수'].loc[loc] = str(read_guild_list2[naming[data1]+'점수'].loc[loc]) + "점 -> " + str(m)
                read_guild_list2[naming[data1]+'점수'].loc[loc+1] = str(read_guild_list2[naming[data1]+'점수'].loc[loc+1]) + "점 -> " + str(u)
                read_guild_list2[naming[data1]+'점수'].loc[loc+2] = str(read_guild_list2[naming[data1]+'점수'].loc[loc+2]) + "점 -> " + str(f)

            '''
            파일 저장
            '''
            new_base_dir3 = './'
            new_file_name3 = '분석'+date+'.xlsx'
            new_file_dir3 = os.path.join(new_base_dir3, new_file_name3)
            l = []

            for name in read_guild_list2.T.index.values:
                l.append(name)

            col1=read_guild_list2.T.columns[-6:-1].to_list()
            col2=read_guild_list2.T.columns[-1:].to_list()        
            col3=read_guild_list2.T.columns[:-6].to_list()
            new_col=col1+col2+col3

            read_guild_list2.T[new_col].to_excel(new_file_dir3,
                                                 na_rep='NaN',
                                                 header=False,
                                                 index=False,
                                                 startrow=0,
                                                 startcol=0)          

            '''
            엑셀파일 분석
            '''
            wb = openpyxl.load_workbook(new_file_dir3)
            sheet = wb.active

            row_count = sheet.max_row
            col_count = sheet.max_column

            '''
            너비 설정
            '''
            width = 15
            for i in range(65, 91):
                sheet.column_dimensions[chr(i)].width=width

            '''
            전체 색칠하기
            '''
            for x in range(1, row_count):
                for y in range(1, col_count+1):
                    c = sheet.cell(row=x, column=y)
                    c.fill = PatternFill(start_color="ccffff", end_color="AAAA31", fill_type="solid")

            '''
            F행 따로 처리
            '''
            sheet.column_dimensions['F'].width=4
            for x in range(1, row_count):
                c = sheet.cell(row=x, column=6)
                c.fill = PatternFill(start_color="123456", end_color="123456", fill_type="solid")

            '''
            못쓰는 사람 처리
            '''
            for row in range(1, row_count):
                if sheet[f"E{row}"].value == "너 못써~" :        
                    for col in range(7, col_count):            
                        if sheet.cell(row=row, column=col).value == ' ':
                            break
                        else:
                            sheet.cell(row=row, column=col).fill = PatternFill(start_color="c12341", end_color="c23411", fill_type="solid")
                            sheet.cell(row=row+1, column=col).fill = PatternFill(start_color="c12341", end_color="c23411", fill_type="solid")

            wb.save(new_file_dir3)

            self.progress.insert(self.text_index, 'analysis fin\n')
            self.text_index = self.text_index + 1.0
            self.main_window.update()        
        
    '''
    종료
    '''
    def exit(self):
        self.main_window.destroy()
        
UI()
